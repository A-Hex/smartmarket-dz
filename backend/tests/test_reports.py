# backend/tests/test_reports.py
"""Unit tests for the PDF/XLSX report generators (pure functions, no DB/HTTP)."""
from openpyxl import load_workbook

from app.services.reports.pdf_report import build_executive_pdf
from app.services.reports.xlsx_report import build_raw_results_xlsx


def test_pdf_report_generates_valid_bytes_with_full_data():
    kpi_result = {"kpis": [{"kpi_type": "cltv", "value": 500.0, "status": "computed"}]}
    validation_result = {
        "normality": {"verdict": "pass"}, "heteroscedasticity": {"verdict": "fail"},
        "autocorrelation": {"verdict": "pass"}, "multicollinearity": {"verdict": "warn"},
        "influence": {"verdict": "pass"}, "overall_verdict": "fail",
    }
    forecast_result = {
        "best_model": "ets", "horizon": 14,
        "history": {"actual": [100 + i for i in range(50)]},
        "forecast": {"point": [151, 152, 153]},
    }
    decisions = [
        {"priority": "high", "title": "Test", "description": "desc", "recommended_action": "do X"},
    ]
    pdf_bytes = build_executive_pdf("Demo Co", "sales.csv", kpi_result, validation_result, forecast_result, decisions)
    assert pdf_bytes.startswith(b"%PDF")
    assert len(pdf_bytes) > 500


def test_pdf_report_handles_missing_data_gracefully():
    pdf_bytes = build_executive_pdf("Demo Co", "sales.csv", None, None, None, [])
    assert pdf_bytes.startswith(b"%PDF")


def test_xlsx_report_has_all_seven_sheets():
    xlsx_bytes = build_raw_results_xlsx(dataset_name="sales.csv")
    wb = load_workbook_from_bytes(xlsx_bytes)
    assert set(wb.sheetnames) == {
        "Descriptif", "Régression", "Validation", "Prévision", "Segments", "KPI", "Décisions"
    }


def test_xlsx_report_descriptive_sheet_uses_real_field_names():
    # Mirrors the actual DescriptiveResult/NumericColumnStats schema (column/missing, not name/null_count)
    # so this test would have caught the key mismatch a shallow fabricated-payload test missed.
    descriptive = {
        "numeric_columns": [
            {"column": "sales", "count": 98, "missing": 2, "mean": 100.0, "median": 95.0,
             "std": 10.0, "min": 1.0, "max": 500.0}
        ]
    }
    xlsx_bytes = build_raw_results_xlsx(dataset_name="sales.csv", descriptive=descriptive)
    wb = load_workbook_from_bytes(xlsx_bytes)
    ws = wb["Descriptif"]
    rows = list(ws.iter_rows(values_only=True))
    assert any(row and row[0] == "sales" for row in rows)


def test_xlsx_report_regression_sheet_has_coefficient_rows():
    regression = {
        "formula": "y ~ x", "r_squared": 0.8, "adj_r_squared": 0.79,
        "coefficients": [
            {"term": "x", "coefficient": 1.2, "std_error": 0.1, "t_stat": 12.0, "p_value": 0.0001,
             "ci_lower": 1.0, "ci_upper": 1.4, "significant": True}
        ],
    }
    xlsx_bytes = build_raw_results_xlsx(dataset_name="sales.csv", regression=regression)
    wb = load_workbook_from_bytes(xlsx_bytes)
    ws = wb["Régression"]
    rows = list(ws.iter_rows(values_only=True))
    assert any(row and row[0] == "x" for row in rows)


def load_workbook_from_bytes(data: bytes):
    import io
    return load_workbook(io.BytesIO(data))
