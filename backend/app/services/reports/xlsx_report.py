# backend/app/services/reports/xlsx_report.py
"""
Raw-results XLSX workbook generator (openpyxl).

One sheet per analysis (descriptive, regression, validation, forecast,
segments, KPIs, decisions), per section 13. Pure-ish function: takes
already-fetched result dicts and returns the workbook as bytes.
"""
import io
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_table(ws: Worksheet, headers: list, rows: list) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)


def build_raw_results_xlsx(
    dataset_name: str,
    descriptive: Optional[dict] = None,
    regression: Optional[dict] = None,
    validation: Optional[dict] = None,
    forecast: Optional[dict] = None,
    segmentation: Optional[dict] = None,
    kpi: Optional[dict] = None,
    decisions: Optional[list] = None,
) -> bytes:
    """Render the raw-results XLSX workbook and return it as raw bytes."""
    wb = Workbook()
    wb.remove(wb.active)

    # --- Descriptive ---------------------------------------------------------
    ws = wb.create_sheet("Descriptif")
    if descriptive and descriptive.get("numeric_columns"):
        _write_table(
            ws,
            ["Colonne", "Moyenne", "Médiane", "Écart-type", "Min", "Max", "Manquants"],
            [
                [c["column"], c.get("mean"), c.get("median"), c.get("std"), c.get("min"), c.get("max"), c.get("missing")]
                for c in descriptive["numeric_columns"]
            ],
        )
    else:
        ws.append(["Aucune statistique descriptive disponible."])

    # --- Regression ----------------------------------------------------------
    ws = wb.create_sheet("Régression")
    if regression and regression.get("coefficients"):
        ws.append([f"Formule : {regression.get('formula', '')}"])
        ws.append([f"R² = {regression.get('r_squared'):.4f}  |  R² ajusté = {regression.get('adj_r_squared'):.4f}"])
        ws.append([])
        _write_table(
            ws,
            ["Variable", "Coefficient", "Erreur std", "t", "p-value", "IC 95% bas", "IC 95% haut", "Significatif"],
            [
                [c["term"], c["coefficient"], c["std_error"], c["t_stat"], c["p_value"],
                 c["ci_lower"], c["ci_upper"], "Oui" if c["significant"] else "Non"]
                for c in regression["coefficients"]
            ],
        )
    else:
        ws.append(["Aucune régression disponible."])

    # --- Validation ------------------------------------------------------------
    ws = wb.create_sheet("Validation")
    if validation:
        rows = []
        for label, key in [
            ("Normalité (Shapiro/Jarque-Bera)", "normality"),
            ("Hétéroscédasticité (Breusch-Pagan)", "heteroscedasticity"),
            ("Autocorrélation (Durbin-Watson)", "autocorrelation"),
            ("Points influents (Cook)", "influence"),
        ]:
            r = validation.get(key) or {}
            rows.append([label, r.get("statistic"), r.get("p_value"), r.get("verdict", "").upper(), r.get("meaning", "")])
        _write_table(ws, ["Test", "Statistique", "p-value", "Verdict", "Signification"], rows)
        ws.append([])
        ws.append(["VIF par variable"])
        vif_rows = [[v["feature"], v["vif"], v["verdict"].upper()] for v in (validation.get("multicollinearity") or {}).get("vif", [])]
        for r in vif_rows:
            ws.append(r)
        ws.append([])
        ws.append([f"Verdict global : {validation.get('overall_verdict', '').upper()}"])
        for rem in validation.get("remediation", []):
            ws.append([f"- {rem}"])
    else:
        ws.append(["Aucune validation disponible."])

    # --- Forecast --------------------------------------------------------------
    ws = wb.create_sheet("Prévision")
    if forecast:
        ws.append([f"Meilleur modèle : {forecast.get('best_model', '').upper()}  |  Horizon : {forecast.get('horizon')} jours"])
        arima = forecast.get("arima_metrics", {})
        ets = forecast.get("ets_metrics", {})
        ws.append([f"ARIMA — MAE={arima.get('mae'):.2f}  RMSE={arima.get('rmse'):.2f}  MAPE={arima.get('mape'):.1f}%"])
        ws.append([f"ETS — MAE={ets.get('mae'):.2f}  RMSE={ets.get('rmse'):.2f}  MAPE={ets.get('mape'):.1f}%"])
        ws.append([])
        fc = forecast.get("forecast", {})
        _write_table(
            ws,
            ["Date", "Prévision", "IC 80% bas", "IC 80% haut", "IC 95% bas", "IC 95% haut"],
            list(zip(
                fc.get("dates", []), fc.get("point", []),
                fc.get("ci_lower_80", []), fc.get("ci_upper_80", []),
                fc.get("ci_lower_95", []), fc.get("ci_upper_95", []),
            )),
        )
    else:
        ws.append(["Aucune prévision disponible."])

    # --- Segments --------------------------------------------------------------
    ws = wb.create_sheet("Segments")
    if segmentation and segmentation.get("clusters"):
        ws.append([f"Algorithme : {segmentation.get('algorithm')}  |  Silhouette : {segmentation.get('silhouette')}"])
        ws.append([])
        _write_table(
            ws,
            ["Cluster", "Nom", "Taille", "Part (%)"],
            [[c["cluster"], c["name"], c["size"], round(c["share"] * 100, 1)] for c in segmentation["clusters"]],
        )
    else:
        ws.append(["Aucune segmentation disponible."])

    # --- KPIs --------------------------------------------------------------------
    ws = wb.create_sheet("KPI")
    if kpi and kpi.get("kpis"):
        _write_table(
            ws,
            ["KPI", "Statut", "Valeur", "Formule"],
            [[k["kpi_type"].upper(), k["status"], k.get("value"), k.get("formula")] for k in kpi["kpis"]],
        )
    else:
        ws.append(["Aucun KPI disponible."])

    # --- Decisions -----------------------------------------------------------------
    ws = wb.create_sheet("Décisions")
    if decisions:
        _write_table(
            ws,
            ["Priorité", "Catégorie", "Titre", "Description", "Action recommandée", "Confiance"],
            [
                [d.get("priority", "").upper(), d.get("category"), d.get("title"), d.get("description"),
                 d.get("recommended_action"), d.get("confidence")]
                for d in decisions
            ],
        )
    else:
        ws.append(["Aucune décision disponible."])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
